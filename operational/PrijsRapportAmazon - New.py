# -*- coding: utf-8 -*-
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple, Optional
import re
import csv
import sys
import platform

import pandas as pd
import openpyxl

# ------------------------------------------------------------
# Optionele override (laat op None voor autodetectie)
# ------------------------------------------------------------
FORCE_YEAR: Optional[int] = 2025          # bijv. 2025
FORCE_MONTH_ABBR: Optional[str] = 'Oct'   # bijv. "Oct" of "Okt"

# ------------------------------------------------------------
# 0) Helpers voor paden (Windows → WSL) en instellingen
# ------------------------------------------------------------
def _win_to_wsl(p: str) -> str:
    if not isinstance(p, str):
        p = str(p)
    p_norm = p.replace("\\", "/")
    if platform.system().lower() == "linux" and p_norm[1:3] == ":/":
        drive = p_norm[0].lower()
        return f"/mnt/{drive}{p_norm[2:]}"
    if platform.system().lower() == "linux" and re.match(r"^[A-Za-z]:/", p_norm):
        drive = p_norm[0].lower()
        return f"/mnt/{drive}{p_norm[2:]}"
    if platform.system().lower() == "linux" and p_norm.startswith("C:/"):
        return "/mnt/c" + p_norm[2:]
    return p_norm

# --- USER SETTINGS ---
ROOT = Path(_win_to_wsl(
    r"C:\Users\semja\Platform Control\Platform Control - Overzichten - Documenten\General\02. Partners\10. Zebra & Friends\04. Analyses\AmazonWinstrapportage"
))

COUNTRY_DIRS: Dict[str, Path] = {
    "DE": ROOT / "DE",
    "FR": ROOT / "FR",
    # "ES": ROOT / "ES",
    # "IT": ROOT / "IT",
    # "NL": ROOT / "NL",
    # "BE": ROOT / "BE",
    # "UK": ROOT / "UK",
    # "SE": ROOT / "SE",
    # "IE": ROOT / "IE",
    "PL": ROOT / "PL"
}

# Gebruik liefst het geüpdatete bestand dat we net maakten
TRANSLATION_WB = Path(
    r"/home/semja/github/personal/PC_amazon_API/operational/Payments report link vertalingen FIXED.xlsx"
)

OUTPUT_FILE = ROOT / "invoicedata.xlsx"

# ------------------------------------------------------------
# 1) Constanten & mapping
# ------------------------------------------------------------
MONTH_ABBR_TO_NUM = {
    "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
    "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Okt": 10, "Nov": 11, "Dec": 12
}

# Engelse header-normalisatie (workbook hanteert deze namen)
ENGLISH_NORMALIZATION = {
    "giftwrap credits tax": "gift wrap credits tax",
    "other transaction fees": "other transactions fees",
}

FINAL_COLS: List[str] = [
    "country", "date/time", "settlement id", "type", "order id", "sku",
    "description", "quantity", "marketplace", "fulfilment", "order city",
    "order state", "order postal", "product sales", "product sales tax",
    "postage credits", "shipping credits tax", "gift wrap credits",
    "gift wrap credits tax", "promotional rebates", "promotional rebates tax",
    "marketplace withheld tax", "selling fees", "fba fees",
    "other transactions fees", "other", "total",
]
MONEY_COLS = [c for c in FINAL_COLS if c not in {
    "country", "date/time", "settlement id", "type", "order id", "sku",
    "description", "quantity", "marketplace", "fulfilment",
    "order city", "order state", "order postal",
}]

# Bestandsnaam-detectie
FNAME_RE = re.compile(
    r"^(?P<year>20\d{2})(?P<mon>Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Okt|Nov|Dec)MonthlyTransaction"
    r"(?:\s*\(\d+\))?\.(?P<ext>csv|CSV|xlsx|XLSX)$"
)

_EU_NBSP = "\u202f"

# ------------------------------------------------------------
# 1b) Valuta → EUR
# ------------------------------------------------------------
CURRENCY_BY_COUNTRY: Dict[str, str] = {
    "SE": "SEK",
    "PL": "PLN",
    "UK": "GBP",
    # overige: DE/FR/ES/IT/NL/BE/IE -> EUR
}

# 1 valuta-eenheid = X EUR  (vul per run in)
FX_TO_EUR: Dict[str, float] = {
    "EUR": 1.0,
    # ECB: 1 EUR = 10.9865 SEK  → 1 SEK = 1/10.9865 = 0.0910208 EUR
    "SEK": 0.091021,
    # ECB: 1 EUR = 4.2570 PLN   → 1 PLN = 1/4.2570  = 0.2349072 EUR
    "PLN": 0.234907,
    # ECB: 1 EUR = 0.8795 GBP   → 1 GBP = 1/0.8795  = 1.1370097 EUR
    "GBP": 1.137010,
}

CONVERT_ALL_MONEY_COLS: bool = True

# ------------------------------------------------------------
# 2) Vertalingen inlezen (kolommen + betalingstypen)
# ------------------------------------------------------------
def build_translation_dicts(wb_path: Path) -> tuple[Dict[str, str], Dict[str, str]]:
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    ws_h = wb['Translated Column Headers']
    ws_p = wb['Translated Type of Payment']

    # headers: row 2 is Engels; vanaf kolom 2
    eng_hdr = [(c, (ws_h.cell(2, c).value or '').strip())
               for c in range(2, ws_h.max_column + 1)
               if (ws_h.cell(2, c).value or '').strip()]
    col_map: Dict[str, str] = {}
    r = 3
    while True:
        if all((ws_h.cell(r, c).value in (None, '') for c, _ in eng_hdr)):
            break
        for c, eng in eng_hdr:
            local = (ws_h.cell(r, c).value or '').strip()
            if local:
                col_map[local.lower()] = eng
        r += 1

    # types: skip kolom 1 (landcode), Engels in row 2 vanaf kolom 2
    eng_types = [(c, (ws_p.cell(2, c).value or '').strip())
                 for c in range(2, ws_p.max_column + 1)
                 if (ws_p.cell(2, c).value or '').strip()]
    pay_map: Dict[str, str] = {}
    r = 3
    while True:
        if all((ws_p.cell(r, c).value in (None, '') for c, _ in eng_types)):
            break
        for c, eng in eng_types:
            local = (ws_p.cell(r, c).value or '').strip()
            if local:
                pay_map[local.lower()] = eng
        r += 1

    return col_map, pay_map

COL_MAP, PAY_MAP = build_translation_dicts(TRANSLATION_WB)

# ------------------------------------------------------------
# 3) Utilities parsing/formatting
# ------------------------------------------------------------
def format_eu(amount: float) -> str:
    sign = "- " if amount < 0 else ""
    amt = abs(amount)
    parts = f"{amt:,.2f}".replace(",", " ").replace(".", ",")
    return f"{sign}€ {parts}"

# Nieuwe robuuste geldparser
DECIMAL_CHARS = set("0123456789")
def parse_money_smart(text: str) -> float:
    s = str(text or "").strip().replace("\u202f","").replace("€","").replace("£","").replace("PLN","").replace("SEK","").replace("zł","").replace(" ","")
    if not s: return 0.0
    if "," in s and "." in s:
        s = s.replace(".", "") if s.rfind(",") > s.rfind(".") else s.replace(",", "")
        s = s.replace(",", ".")
    elif "," in s:
        s = s.replace(".", "").replace(",", ".")
    try: return float(s)
    except: return 0.0

def reformat_date(series: pd.Series, month_num: int, german: bool) -> pd.Series:
    if german:
        pattern = r"(\d{1,2})\.(\d{1,2})\.(\d{4})"
        func = lambda m: f"{int(m.group(1)):02d}-{month_num:02d}-{m.group(3)}"
    else:
        pattern = r"(\d{1,2})\s+\S+\s+(\d{4})"
        func = lambda m: f"{int(m.group(1)):02d}-{month_num:02d}-{m.group(2)}"
    return series.astype(str).str.replace(pattern, func, regex=True)

# ------------------------------------------------------------
# 4) Bestanden vinden
# ------------------------------------------------------------
@dataclass(frozen=True)
class CandidateFile:
    country: str
    path: Path
    year: int
    month_num: int
    ext: str

def scan_country_dir(country: str, dir_path: Path | str) -> List[CandidateFile]:
    p = Path(dir_path)
    if not p.exists():
        return []
    out: List[CandidateFile] = []
    for f in p.iterdir():
        if not f.is_file():
            continue
        m = FNAME_RE.match(f.name)
        if not m:
            continue
        year = int(m.group("year"))
        mon_abbr = m.group("mon")
        month_num = MONTH_ABBR_TO_NUM[mon_abbr]
        ext = m.group("ext").lower()
        out.append(CandidateFile(country, f, year, month_num, ext))
    return out

def pick_target_period(cands: List[CandidateFile]) -> Optional[Tuple[int, int]]:
    if not cands:
        return None
    return max(((c.year, c.month_num) for c in cands))

def pick_files_for_period(cands: List[CandidateFile], year: int, month_num: int) -> Dict[str, CandidateFile]:
    per_country: Dict[str, List[CandidateFile]] = {}
    for c in cands:
        if c.year == year and c.month_num == month_num:
            per_country.setdefault(c.country, []).append(c)
    chosen: Dict[str, CandidateFile] = {}
    for country, lst in per_country.items():
        lst_sorted = sorted(lst, key=lambda x: x.path.stat().st_mtime, reverse=True)
        chosen[country] = lst_sorted[0]
    return chosen

# ------------------------------------------------------------
# 5) Lezen + normaliseren (incl. EUR conversie)
# ------------------------------------------------------------
def read_monthly_transaction(cf: CandidateFile) -> pd.DataFrame:
    print(f"📥  {cf.country}: {cf.path.name}")

    if cf.ext == "csv":
        df = pd.read_csv(cf.path, skiprows=7, dtype=str, quoting=csv.QUOTE_MINIMAL).fillna("")
    elif cf.ext == "xlsx":
        df = pd.read_excel(cf.path, sheet_name=0, dtype=str).fillna("")
    else:
        raise ValueError(f"Niet-ondersteunde extensie: {cf.ext}")

    # Kolommen vertalen en normaliseren
    df = df.rename(columns={c: COL_MAP.get(str(c).lower(), c) for c in df.columns})
    df = df.rename(columns=lambda c: ENGLISH_NORMALIZATION.get(c, c))

    # Zorg dat alle verwachte kolommen bestaan
    for col in FINAL_COLS:
        if col not in df.columns:
            df[col] = ""

    # Type vertalen
    df["type"] = df["type"].map(lambda x: PAY_MAP.get(str(x).lower(), x))

    # Land + datum
    df["country"] = cf.country
    df["date/time"] = reformat_date(df["date/time"], cf.month_num, german=(cf.country == "DE"))

    # 'Transfer' uitsluiten
    df = df[df["type"] != "Transfer"].copy()

    # Money → helper floats (lokale valuta) + lokale EU-notatie voor output
    for col in MONEY_COLS:
        df[f"_{col}_float"] = df[col].apply(parse_money_smart)
        df[col] = df[f"_{col}_float"].map(lambda v: f"{v:,.2f}".replace(",", " ").replace(".", ","))

    # ---- EUR conversie ----
    currency = CURRENCY_BY_COUNTRY.get(cf.country, "EUR")
    rate = FX_TO_EUR.get(currency)
    if rate is None:
        raise RuntimeError(f"FX-koers ontbreekt voor valuta '{currency}' (land {cf.country}). Vul FX_TO_EUR in.")

    df["_fx_currency"] = currency
    df["_fx_rate_to_eur"] = rate

    for col in MONEY_COLS:
        df[f"_{col}_eur"] = df[f"_{col}_float"] * rate
        if CONVERT_ALL_MONEY_COLS:
            df[f"{col} (EUR)"] = df[f"_{col}_eur"].apply(format_eu)

    # --- sanity check: parsed floats vs. zichtbaar geformatteerde strings ---
    for m in ("product sales", "selling fees", "fba fees", "total"):
        src_sum = df[f"_{m}_float"].sum()
        out_sum = df[m].apply(parse_money_smart).sum()
        if abs(src_sum - out_sum) > 1e-6:
            raise AssertionError(
                f"Mismatch in {cf.country} ({cf.path.name}) for '{m}': "
                f"{src_sum:.6f} vs {out_sum:.6f}"
            )

    return df

# ------------------------------------------------------------
# 6) Main
# ------------------------------------------------------------
def main() -> None:
    ROOT.mkdir(parents=True, exist_ok=True)

    # Scan alle landen
    all_cands: List[CandidateFile] = []
    for cc, d in COUNTRY_DIRS.items():
        all_cands.extend(scan_country_dir(cc, d))

    if not all_cands:
        print("Geen geschikte bestanden gevonden in COUNTRY_DIRS.")
        sys.exit(1)

    # Periode bepalen
    if FORCE_YEAR and FORCE_MONTH_ABBR:
        if FORCE_MONTH_ABBR not in MONTH_ABBR_TO_NUM:
            print(f"Onbekende maandafkorting: {FORCE_MONTH_ABBR}. "
                  f"Gebruik een van: {', '.join(MONTH_ABBR_TO_NUM.keys())}")
            sys.exit(1)
        target = (FORCE_YEAR, MONTH_ABBR_TO_NUM[FORCE_MONTH_ABBR])
    else:
        target = pick_target_period(all_cands)

    if not target:
        print("Geen periode gedetecteerd.")
        sys.exit(1)

    year, month_num = target
    print(f"ℹ️  Geselecteerde periode: {year}-{month_num:02d}")

    chosen = pick_files_for_period(all_cands, year, month_num)
    if not chosen:
        print("Geen bestanden per land gevonden voor de geselecteerde periode.")
        sys.exit(1)

    # Lees en combineer
    frames: List[pd.DataFrame] = []
    for cc in sorted(chosen.keys()):
        try:
            frames.append(read_monthly_transaction(chosen[cc]))
        except Exception as e:
            print(f"⚠️  {cc}: overslaan door leesfout: {e}")

    if not frames:
        print("Geen enkele marketplace succesvol verwerkt.")
        sys.exit(1)

    combined = pd.concat(frames, ignore_index=True)

    # Outputkolommen
    eur_cols = [f"{c} (EUR)" for c in MONEY_COLS] if CONVERT_ALL_MONEY_COLS else []
    meta_cols = ["_fx_currency", "_fx_rate_to_eur"]
    output_cols = FINAL_COLS + eur_cols + meta_cols

    # Schrijf output
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    combined.to_excel(OUTPUT_FILE, index=False, columns=output_cols)
    print(f"\n✅  Geconsolideerd rapport geschreven naar:\n    {OUTPUT_FILE}")

    # Reconciliatie
    metrics = {
        "product sales":   "_product sales_float",
        "selling fees":    "_selling fees_float",
        "fba fees":        "_fba fees_float",
        "total":           "_total_float",
    }
    metrics_eur = {
        "product sales (EUR)": "_product sales_eur",
        "selling fees (EUR)":  "_selling fees_eur",
        "fba fees (EUR)":      "_fba fees_eur",
        "total (EUR)":         "_total_eur",
    }

    print("\nReconciliation (alle landen gecombineerd):")
    print(f"{'Metric':<22}{'Source CSV/XLSX':>20}{'Output XLSX':>18}{'Match?':>8}")
    for m, helper_col in metrics.items():
        src_sum = combined[helper_col].sum()
        out_sum = combined[m].apply(parse_money_smart).sum()
        ok = abs(src_sum - out_sum) < 1e-6
        print(f"{m:<22}{format_eu(src_sum):>20}{format_eu(out_sum):>18}{'✅' if ok else '❌':>8}")

    print("\nEUR-sommen (na FX):")
    for m, helper_col in metrics_eur.items():
        print(f"{m:<22}{format_eu(combined[helper_col].sum()):>20}")

if __name__ == "__main__":
    main()