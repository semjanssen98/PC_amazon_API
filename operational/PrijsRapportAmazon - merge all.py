# -*- coding: utf-8 -*-
"""Amazon payment reports → outputdata.xlsx
================================================

Scans every *.csv in COUNTRY_DIRS, translates/merges, writes a single
workbook (outputdata.xlsx) and prints a reconciliation table.
"""
from __future__ import annotations

from pathlib import Path
from typing import Dict, List
import re, csv, unicodedata

import pandas as pd
import openpyxl

# ---------------------------------------------------------------------------
# 1. USER SETTINGS
# ---------------------------------------------------------------------------

COUNTRY_DIRS: Dict[str, str] = {
    "DE": r"C:\Users\semja\Platform Control\Platform Control - Overzichten - General\02. Partners\03. Animal Platform Control\09. Analyses\AmazonWinstrapportage\DE",
    "FR": r"C:\Users\semja\Platform Control\Platform Control - Overzichten - General\02. Partners\03. Animal Platform Control\09. Analyses\AmazonWinstrapportage\FR",
    "ES": r"C:\Users\semja\Platform Control\Platform Control - Overzichten - General\02. Partners\03. Animal Platform Control\09. Analyses\AmazonWinstrapportage\ES",
    "IT": r"C:\Users\semja\Platform Control\Platform Control - Overzichten - General\02. Partners\03. Animal Platform Control\09. Analyses\AmazonWinstrapportage\IT",
    "NL": r"C:\Users\semja\Platform Control\Platform Control - Overzichten - General\02. Partners\03. Animal Platform Control\09. Analyses\AmazonWinstrapportage\NL",
    "BE": r"C:\Users\semja\Platform Control\Platform Control - Overzichten - General\02. Partners\03. Animal Platform Control\09. Analyses\AmazonWinstrapportage\BE"

}

ROOT_OUTPUT = Path(
    r"C:\Users\semja\Platform Control\Platform Control - Overzichten - General\02. Partners\03. Animal Platform Control\09. Analyses\AmazonWinstrapportage"
)
TRANSLATION_WB = Path(
    r"C:\Users\semja\Platform Control\Platform Control - Owners - General\Sem (niet verwijderen)\1. Documenten\1. Back-up\Gitbackup\home\semja\projects\Operational\Prijsrapporten_amazon\Payments report link vertalingen.xlsx"
)
OUTPUT_FILE = ROOT_OUTPUT / "outputdata.xlsx"

# ---------------------------------------------------------------------------
# 2. CONSTANTS (normally leave alone)
# ---------------------------------------------------------------------------

FINAL_COLS = [
    "country", "date/time", "settlement id", "type", "order id", "sku",
    "description", "quantity", "marketplace", "fulfilment", "order city",
    "order state", "order postal", "product sales", "product sales tax",
    "postage credits", "shipping credits tax", "gift wrap credits",
    "gift wrap credits tax", "promotional rebates", "promotional rebates tax",
    "marketplace withheld tax", "selling fees", "fba fees",
    "other transactions fees", "other", "total",
]

MONEY_COLS = [
    c for c in FINAL_COLS
    if c not in {
        "country", "date/time", "settlement id", "type", "order id", "sku",
        "description", "quantity", "marketplace", "fulfilment",
        "order city", "order state", "order postal",
    }
]

_EU_NBSP = "\u202f"           # narrow no-break space used by Amazon EU

# ---------------------------------------------------------------------------
# 2a.  Month names & auto-generated abbreviations
# ---------------------------------------------------------------------------

_MONTH_TABLE = {
    "EN": ["January","February","March","April","May","June",
           "July","August","September","October","November","December"],
    "PL": ["Styczeń","Luty","Marzec","Kwiecień","Maj","Czerwiec",
           "Lipiec","Sierpień","Wrzesień","Październik","Listopad","Grudzień"],
    "ES": ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
           "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"],
    "IT": ["Gennaio","Febbraio","Marzo","Aprile","Maggio","Giugno",
           "Luglio","Agosto","Settembre","Ottobre","Novembre","Dicembre"],
    "FR": ["Janvier","Février","Mars","Avril","Mai","Juin",
           "Juillet","Août","Septembre","Octobre","Novembre","Décembre"],
    "DE": ["Januar","Februar","März","April","Mai","Juni",
           "Juli","August","September","Oktober","November","Dezember"],
    "NL": ["Januari","Februari","Maart","April","Mei","Juni",
           "Juli","Augustus","September","Oktober","November","December"],
}

def _strip_accents(s: str) -> str:
    return "".join(
        ch for ch in unicodedata.normalize("NFD", s)
        if unicodedata.category(ch) != "Mn"
    )

MONTH_MAP: dict[str, int] = {}
for names in _MONTH_TABLE.values():
    for idx, full in enumerate(names, 1):
        base = _strip_accents(full).lower()
        MONTH_MAP[base] = idx                 # full name, accent-stripped
        MONTH_MAP[full.lower()] = idx         # exact spelling
        for n in (3, 4):                      # abbreviations
            if len(base) >= n:
                abbr = base[:n]
                MONTH_MAP[abbr] = idx
                MONTH_MAP[abbr + "."] = idx

# extras not generated automatically
MONTH_MAP.update({"mrt": 3, "mei": 5})

_NUMERIC_RE = re.compile(r"^(\d{1,2})\.(\d{1,2})\.(\d{4})")
_TEXT_RE    = re.compile(r"^(\d{1,2})\s+([A-Za-zÀ-ÿ\.]+)\s+(\d{4})")

# ---------------------------------------------------------------------------
# 3.  Translation workbook → dicts
# ---------------------------------------------------------------------------

def build_translation_dicts(wb_path: Path):
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    ws_h, ws_p = wb[wb.sheetnames[0]], wb[wb.sheetnames[1]]

    # header translations ----------------------------------------------------
    eng_hdr = [(ws_h.cell(2, c).value or "").strip() for c in range(2, 29)]
    col_map: dict[str, str] = {}
    row = 3
    while any(ws_h.cell(row, c).value not in (None, "") for c in range(2, 29)):
        for c, eng in zip(range(2, 29), eng_hdr):
            local = (ws_h.cell(row, c).value or "").strip()
            if local:
                col_map[local.lower()] = eng
        row += 1

    # payment-type translations ---------------------------------------------
    eng_types, c = [], 1
    while (val := ws_p.cell(2, c).value):
        eng_types.append(str(val).strip()); c += 1

    pay_map: dict[str, str] = {}
    row = 3
    while any(ws_p.cell(row, i).value not in (None, "") for i in range(1, len(eng_types)+1)):
        for idx, eng in enumerate(eng_types, 1):
            local = (ws_p.cell(row, idx).value or "").strip()
            if local:
                pay_map[local.lower()] = eng
        row += 1

    return col_map, pay_map

COL_MAP, PAY_MAP = build_translation_dicts(TRANSLATION_WB)

# ---------------------------------------------------------------------------
# 4.  Helpers
# ---------------------------------------------------------------------------

parse_num = lambda s: 0.0 if not s else float(
    s.replace(_EU_NBSP, "").replace(" ", "").replace(".", "").replace(",", ".")
)
fmt_eu = lambda v: ("- " if v < 0 else "") + f"€ {abs(v):,.2f}".replace(",", " ").replace(".", ",")

def norm_date(text: str) -> str:
    """Return dd-mm-yyyy for any EU Amazon date string."""
    text = text.strip()

    if (m := _NUMERIC_RE.match(text)):                 # 10.12.2024
        d, m_, y = map(int, m.groups())
        return f"{d:02d}-{m_:02d}-{y}"

    if (m := _TEXT_RE.match(text)):                    # 15 abr 2025, 1 févr. 2025 …
        d, mw_raw, y = m.groups()
        key = _strip_accents(mw_raw.lower().rstrip("."))
        if (mn := MONTH_MAP.get(key)):
            return f"{int(d):02d}-{mn:02d}-{y}"

    return text                                        # fallback: unchanged

# ---------------------------------------------------------------------------
# 5.  File → DataFrame
# ---------------------------------------------------------------------------

def process_file(csv_path: Path, cc: str) -> pd.DataFrame:
    df = (
        pd.read_csv(csv_path, skiprows=7, dtype=str, quoting=csv.QUOTE_MINIMAL)
        .fillna("")
    )

    # rename → English
    df = df.rename(columns={c: COL_MAP[c.lower()] for c in df.columns if c.lower() in COL_MAP})

    # ensure expected columns present
    for col in FINAL_COLS:
        if col not in df.columns:
            df[col] = ""

    # translate payment-type, normalise date
    df["type"] = df["type"].map(lambda x: PAY_MAP.get(x.lower(), x))
    df["country"] = cc
    df["date/time"] = df["date/time"].map(norm_date)
    df = df[df["type"] != "Transfer"].copy()

    # money columns → helper floats + EU formatting
    for col in MONEY_COLS:
        df[f"_{col}_f"] = df[col].map(parse_num)
        df[col] = (
            df[col]
              .str.replace(_EU_NBSP, "")
              .str.replace(".", ",", regex=False)
        )

    return df[FINAL_COLS + [f"_{c}_f" for c in MONEY_COLS]]

# ---------------------------------------------------------------------------
# 6.  MAIN
# ---------------------------------------------------------------------------

def main() -> None:
    frames: List[pd.DataFrame] = []

    for cc, dir_str in COUNTRY_DIRS.items():
        folder = Path(dir_str)
        if not folder.exists():
            print(f"⚠ Folder not found: {folder}")
            continue

        csv_files = sorted(folder.glob("*.csv"))
        print(f"📂 {cc} — {len(csv_files)} CSV file(s)")

        for csv_path in csv_files:
            print(f"     → {csv_path.name}", end="")
            df = process_file(csv_path, cc)
            print(f"  ({len(df)} rows)")
            frames.append(df)

    if not frames:
        print("No CSV files found; exiting."); return

    combined = pd.concat(frames, ignore_index=True)
    combined[FINAL_COLS].to_excel(OUTPUT_FILE, index=False)
    print("\n📦  outputdata.xlsx written to:\n   ", OUTPUT_FILE)

    metrics = {m: f"_{m}_f" for m in ["product sales", "selling fees", "fba fees", "total"]}
    print("\nReconciliation (all countries combined):\n")
    print(f"{'Metric':<20}{'Source CSV':>18}{'Output XLSX':>18}{'Match?':>8}")
    for m, hcol in metrics.items():
        src = combined[hcol].sum()
        out = combined[m].map(parse_num).sum()
        print(f"{m:<20}{fmt_eu(src):>18}{fmt_eu(out):>18}{'✅' if abs(src-out) < 1e-6 else '❌':>8}")

if __name__ == "__main__":
    main()